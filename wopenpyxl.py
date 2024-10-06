import datetime
from copy import copy

import openpyxl as px
from openpyxl.styles import PatternFill
from openpyxl.workbook import child
from openpyxl.worksheet.worksheet import Worksheet

# 2007 Office system ファイル形式の MIME タイプをサーバーで登録する
# https://technet.microsoft.com/ja-jp/library/ee309278(v=office.12).aspx
XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class WOpenpyxl:
    """openpyxlラッパー
    基本はコンストラクタでbookを作ってactive_sheetを操作
    備忘録的なラッパーで、細かい命令はopenpyxl自体を利用する想定
    """

    def __init__(self, path=None, **kwargs):
        """
        :param path: 開くxlsxパス、指定なしで新規作成
        """
        self.__path = None
        self.__book = None
        self.__sheet = None

        if path is None:
            self.__constructor_newbook()
        else:
            self.__constructor_existsbook(path, **kwargs)

    def __constructor_newbook(self):
        """book新規作成"""
        self.__book = px.Workbook()
        self.__sheet = self.__book.active

    def __constructor_existsbook(self, path, **kwargs):
        """book開く"""
        self.__path = path
        self.__book = px.load_workbook(path, **kwargs)
        self.__sheet = self.__book.active

    def close(self):
        self.__book.close()
        del self.__sheet
        del self.__book

    def get_openpyxl_book(self):
        return self.__book

    @property
    def active_sheet(self):
        return self.__sheet

    def set_active_sheet(self, sheet_name):
        self.__sheet = self.__book[sheet_name]

    def get_sheet(self, sheet_name):
        return self.__book[sheet_name]

    def create_sheet(self, sheet_name):
        return self.__book.create_sheet(sheet_name)

    def remove_sheet(self, sheet: Worksheet):
        self.__sheet = self.__book.remove(sheet)

    def remove_sheet(self, sheet_name: str):
        self.__sheet = self.__book.remove(self.__book.get_sheet_by_name(sheet_name))

    def delete_rows(self, idx: int, amount: int = 1):
        self.__sheet.delete_rows(idx, amount)

    def iter_rows(self, min_row=None, max_row=None):
        """指定の行を返す"""
        min_row = min_row or 1
        max_row = max_row or self.max_rows()
        return self.__sheet.iter_rows(min_row=min_row, max_row=max_row)

    def iter_cols(self, min_col=None, max_col=None):
        """指定の列を返す"""
        min_col = min_col or 1
        max_col = max_col or self.max_cols()
        return self.__sheet.iter_cols(min_col=min_col, max_col=max_col)

    def max_rows(self):
        """最大行番号取得
        :return: 最大行番号
        """
        return self.__sheet.max_row

    def max_cols(self):
        """最大列番号取得
        :return: 最大列番号

        Notes
        -----
        低速注意
        """
        count = 0
        for col in self.__sheet.iter_cols():
            count += 1
        return count

    def save_as(self, path):
        """
        名前を付けて保存
        :param path: 保存先パス
        """
        self.__path = path
        self.__book.save(path)

    def save(self):
        """保存"""
        self.__book.save(self.__path)

    def clear(self):
        """シートクリア"""
        # TESTME
        for row in self.__sheet.iter_rows():
            for cell in row:
                cell.value = None

    def clear_fa1(self, fa1):
        """指定範囲クリア"""
        # TESTME
        for row in self.__sheet[fa1]:
            for cell in row:
                cell.value = None

    @property
    def file_path(self):
        """ファイルパス"""
        return self.__path

    @file_path.setter
    def file_path(self, path):
        """ファイルパス"""
        self.__path = path

    @property
    def sheet_name(self):
        """シート名"""
        return self.__sheet.title

    @sheet_name.setter
    def sheet_name(self, title):
        """シート名"""
        # TESTME
        self.__sheet.title = title

    def get_cell(self, row, col):
        """R1C1形式でセル取得"""
        return self.__sheet.cell(row, col)

    def get_value_a1(self, a1):
        """A1形式で値取得"""
        return self.__sheet[a1].value

    def set_value_a1(self, a1, value):
        """A1形式で値設定"""
        self.__sheet[a1] = value

    def get_value(self, row, col):
        """R1C1形式で値取得"""
        return self.__sheet.cell(row, col).value

    def set_value(self, row, col, value):
        """R1C1形式で値設定"""
        self.__sheet.cell(row, col, value)

    def set_fg_color(self, row, col, color_code, pattern_type="solid"):
        """R1C1形式で値設定"""
        fill = PatternFill(patternType=pattern_type, fgColor=color_code)
        self.__sheet.cell(row, col).fill = fill

    def append(self, rows):
        """行データを追記
        :param rows:
            column_title = ["FirstName", "LastName"]
            rows = [
                column_title,
                ["Tarou", "Tanaka"],
                ["Tarou", "Suzuki"],
                ["Tarou", "Uchiayama"],
            ]
        :return:
        """
        for row in rows:
            self.__sheet.append(row)

    def set_wrap_text(self, row, col, flag=True):
        """セル内改行
        See Also
        https://openpyxl.readthedocs.io/en/default/styles.html#cell-styles-and-named-styles
        """
        self.__sheet.cell(row, col).alignment = px.styles.Alignment(wrapText=flag)

    def copy_worksheet(self, source_sheet, new_sheet_name) -> Worksheet:
        new_sheet_name = child.INVALID_TITLE_REGEX.sub("", new_sheet_name)
        new_sheet_name = new_sheet_name[:31]
        new_sheet = self.__book.copy_worksheet(source_sheet)
        new_sheet.title = new_sheet_name
        return new_sheet

    @classmethod
    def make_date(cls, year: int, month: int, day: int):
        """openpyxlで使える日付値生成"""
        return datetime.date(year, month, day)

    @classmethod
    def make_datetime(
        cls,
        year: int,
        month: int,
        day: int,
        hour: int,
        minute: int,
        second: int,
        microsecond: int,
    ):
        """openpyxlで使える日時値生成"""
        return datetime.datetime(year, month, day, hour, minute, second, microsecond)

    def copy_range(
        self,
        template_range_start_a1,
        template_range_end_a1,
        target_range_start_row_num,
        do_copy_value=True,
        do_copy_style=True,
    ):
        """
        :param template_range_start_a1: 書式テンプレート行
        :param template_range_end_a1:
        :param target_range_start_row_num:
        :param do_copy_value:
        :param do_copy_style:
        :return:
        """
        # 空行に次の行の書式をコピーする
        target_range_start = self.__sheet.cell(
            target_range_start_row_num, 1
        )  # コピー先の開始位置
        to_row = target_range_start.row  # コピー先の終了位置
        to_col = target_range_start.col_idx  # コピー先の終了位置
        for template_row_num, template_row in enumerate(
            self.__sheet[template_range_start_a1:template_range_end_a1]
        ):
            if not template_row:
                continue

            for template_col_num, template_col in enumerate(template_row):
                target_cell = self.__sheet.cell(
                    row=to_row + template_row_num, column=to_col + template_col_num
                )

                # cellの値をコピー
                if do_copy_value:
                    target_cell.value = template_col.value

                # cellのスタイルをコピー
                if do_copy_style and template_col.has_style:
                    # _styleはprotectedだけどアクセスするしかない
                    target_cell._style = copy(template_col._style)
